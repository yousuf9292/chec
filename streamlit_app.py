import openpyxl
from openpyxl import Workbook, load_workbook
import streamlit as st
import pandas as pd
import os
import plotly_express as px
import numpy as np
from PIL import Image

# Change the current working directory to H:\Python\Index
#current_dir = os.getcwd()

#print('aaaa'+str(current_dir))

#os.chdir("yousuf9292/app/chec/Index")

# Get the current directory
#current_dir = os.getcwd()

# Get the names of the files in the current directory
#files = [file for file in os.listdir('.') if file.endswith('.xlsx')]

# Workbooks list
lista_wb = ['GP Nr. 05,06,08.xlsx', 'GP Nr. 10-12.xlsx', 'GP Nr. 13-15.xlsx', 'GP Nr. 16-18.xlsx',
            'GP Nr. 19-21.xlsx', 'GP Nr. 22-23.xlsx', 'GP Nr. 24-25.xlsx', 'GP Nr. 26-27.xlsx', 'GP Nr. 28.xlsx',
            'GP Nr. 29-33.xlsx', 'GP Nr. 35,36,38.xlsx', 'lista_nombrearchivos.py']
# Create an empty list to store the new list with excel files
list_xlsx = []

# Iterate over the original list and check if each element contains ".xslx"
for item in lista_wb:
    if '.xlsx' in item:
        list_xlsx.append(item)

# Use the map function to remove the ".xlsx" part from each element in the list
# and create a new list with the modified elements
lista_wb_without_ext = list(map(lambda x: x.replace(".xlsx", ""), lista_wb))

st.set_page_config(page_title="Dashboard Index",
                   page_icon=":bar_chart:",
                   layout="wide"
)

st.sidebar.title("Filter (Tabelle, Line Plot und Index-Vergleich)")
selected_workbook = st.sidebar.selectbox("GP Gruppe:", lista_wb_without_ext)

# Use the selected option to retrieve the corresponding Excel file from the original list
selected_file = lista_wb[lista_wb_without_ext.index(selected_workbook)]
print(selected_file)

# Construct the full path to the file
file_path = 'https://view.officeapps.live.com/op/view.aspx?src=https%3A%2F%2Fraw.githubusercontent.com%2Fyousuf9292%2Fchec%2Fmain%2FIndex%2FGP%2520Nr.%252005%252C06%252C08.xlsx&wdOrigin=BROWSELINK'

# Open the Excel file using the pandas ExcelFile class
excel_file = pd.ExcelFile(file_path,engine='openpyxl')

print(excel_file)

# Get a list of the sheet names in the file
sheet_names = excel_file.sheet_names

# Add a selectbox widget to the sidebar for selecting a sheet
selected_sheet = st.sidebar.selectbox("GP-Nummer wählen", sheet_names)

# Read the selected sheet from the Excel file and create a DataFrame
# with the range A2:N21
df_original = pd.read_excel(excel_file, sheet_name=selected_sheet, usecols="A:N", skiprows=1)

# Drop the last column from the DataFrame
df_original = df_original.drop(df_original.columns[-1], axis=1)

# List columns with float values
lista_columnas_float = ['Jan', 'Feb', 'Mrz', 'Apr', 'Mai', 'Jun', 'Jul', 'Aug', 'Sep', 'Okt', 'Nov', 'Dez']

# Replace the suffix in the first column
df_original[df_original.columns[0]] = df_original[df_original.columns[0]].str.replace(" ...", "")
df_original.replace("-", np.NaN, inplace=True)


df_formatted = df_original.style.format({col: '{:.2f}'.format for col in lista_columnas_float})

# Get the unique values from the first column of the DataFrame
values = df_original[df_original.columns[0]].unique()

headers = df_original.columns.values
meses = headers[1:]

# Add a selectbox widget to the sidebar for choosing one of the values from the first column of the DataFrame
selected_value = st.sidebar.multiselect("Jahr auswählen (Mehrfachauswahl möglich):", values, default="2022")

# Titulo de la pagina
st.title("Dashboard - Erzeugerpreise")
texto_inicio = """
Dieses Dashboard sammelt die Informationen des Erzeugerpreisindex gewerblicher Produkte, um sie visuell und agil darzustellen. 

Die gesammelten Daten werden aus der von DESTATIS bereitgestellten Excel-Datei und in diesem Link veröffentlicht: 
https://www.destatis.de/DE/Themen/Wirtschaft/Preise/Erzeugerpreisindex-gewerbliche-Produkte/_inhalt.html#_u5ovo2o3d.
 
Wenn Sie die Nummer des von Ihnen interessierten Index nicht kennen, finden Sie am unteren Ende der Webseite ein kleines Suchwerkzeug.

Den aktuellen Link des Dashboards wird immer in dem Linkedin Profil des Erstellers gepostet: 
https://www.linkedin.com/in/roberto-s-8a0b81b0/recent-activity/shares/

Das Dashboard wird versucht, aktuell und fehlerfrei zu halten, aber der Benutzer ist für die 
Überprüfung der Richtigkeit der Daten verantwortlich. Alle Angaben ohne Gewähr.
"""

st.markdown(texto_inicio)


st.markdown("---")
sheet_name = selected_sheet
wb = load_workbook(file_path)
# Get a list of all the sheets in the workbook
sheets = wb.worksheets
# Iterate over the sheets
sheet_name_title = []
for sheet in sheets:
    # Check if the sheet's name matches the user input
    if sheet.title == sheet_name:
        # Access the data in the sheet
            if not sheet['D1'].value == None:
                titulo = sheet['D1'].value
            elif not sheet['E1'].value == None:
                titulo = sheet['E1'].value
            else:
                titulo = sheet['F1'].value
            st.title(titulo)




# List to store the tuples of sheet name and cell D1 value
data = []

# Loop through each workbook in the list
for wb in list_xlsx:
    # Open the workbook
    workbook = load_workbook(wb)

    # Loop through each sheet in the workbook
    for sheet in workbook.worksheets:
        # Get the sheet name and cell D1 value
        name = sheet.title
        cell_d1 = sheet['D1'].value

        # Store the sheet name and cell D1 value in a tuple
        # and append the tuple to the list
        data.append((name, cell_d1))

    # Close the workbook
    workbook.close()

# Print the list of tuples
print(data)






# Display the Dataframe
col1, col2 = st.columns(2)
col1.dataframe(df_formatted)


# Transpose the dataframe so that the first row becomes the name for each line
df_transposed = df_original.set_index("Jahr").T

# Create the line plot
fig = px.line(df_transposed, y=selected_value, title="Line-Diagramm aus den ausgewählten Jahren", height=400,
    labels={"variable": "Jahr",  "index": "Monat", "value": "Wert"}
)
fig.update_layout(
    margin={"l": 50, "r": 50, "t": 50, "b": 50},
    legend={"title": "Jahre"}
)
fig.update_xaxes(title="Monat")
fig.update_yaxes(title="Wert")

# Show the plot
col2.plotly_chart(fig)

st.markdown("---")
# Eingabe selector

col1.markdown("---")
col2.markdown("---")


primer_ano = col1.selectbox(
    'Wählen Sie den Jahr',
    values,
    key= 1
)
primer_mes = col1.selectbox(
    'Wählen Sie den Monat',
    meses,
    key= 2
)
segundo_ano = col2.selectbox(
    'Wählen Sie den Jahr',
    values,
    key= 3
)
segundo_mes = col2.selectbox(
    'Wählen Sie den Monat',
    meses,
    key = 4
)

valueelegido = df_original.loc[df_original["Jahr"] == primer_ano, primer_mes].iloc[0]
valueelegido_2 = df_original.loc[df_original["Jahr"] == segundo_ano, segundo_mes].iloc[0]

col1.metric(label="First Value", value=valueelegido)

delta = round((valueelegido_2 / valueelegido - 1) * 100,2)
col2.metric(label="First Value", value=valueelegido_2, delta=f"{delta} %",
    delta_color="inverse")



# Create the search box
texto_buscador = """Bitte geben Sie einen Suchbegriff ein und bestätigen Sie ihn durch Drücken der Enter-Taste.

Beispiele können "Stahl" oder "Kupfer" sein.
"""
st.markdown(texto_buscador)
search_term = st.text_input('')

# Create an empty list to store the matching values
matching_values = []

# Iterate through the tuple of values and check if each value contains the search term
for value in data:
    # Convert the tuple value to a string and use the lower() method to make the search case-insensitive
    if search_term.lower() in str(value).lower():
        matching_values.append(value)

# If there are any matching values, display them
if len(matching_values) > 0:
    st.write('Die folgenden Werte entsprechen Ihrem Suchbegriff:')
    st.write(matching_values)
else:
    st.write(f'Keine Werte gefunden, die dem Suchbegriff entsprechen. "{search_term}"')
st.markdown("---")

#About
st.sidebar.markdown("---")
st.sidebar.title("Ersteller")
imagen = Image.open("H:\\1669069928538NEU.png")
st.sidebar.image(imagen)
st.sidebar.subheader("Roberto SL")
texto_creador = """
Hallo! Ich bin Roberto 
und arbeite derzeit als Kalkulator (Hochbau-Stahlbau). 
Ich bin dabei, meine Programmierfähigkeiten zu entwickeln 
und das ist der Grund, warum wir hier sind.
Ich dachte, dieses Dashboard könnte für andere Leute nützlich sein. 
"""
st.sidebar.text(texto_creador)

texto_final = """
Möchten Sie einen Fehler melden, haben Sie Fragen oder Beschwerden? Meine Kontakt-E-Mail ist rsanzlopes.info.de@gmail.com

Wie bereits zu Beginn dieses Dashboards erwähnt, werden die Ursprungsdaten vom Statistischen Bundesamt erhoben und veröffentlicht. 

Logischerweise wird das Dashboard erst dann aktualisiert, wenn neue Daten in dem hier verwendeten Excel-Format veröffentlicht werden.

Das Dashboard wurde auf gemeinnütziger Basis und als Teil eines Lernprozesses erstellt.

Der Code ist auf Github unter dem folgenden Link veröffentlicht. Jeder Beitrag oder Hinweis zur Verbesserung des Dashboards ist willkommen.
"""

st.markdown(texto_final)
